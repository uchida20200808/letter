#!/usr/bin/env python
# coding: utf-8

# In[41]:


from pathlib import Path
import glob
from openpyxl import load_workbook, Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import pandas as pd
import datetime


# In[42]:


files = glob.glob('*更新リスト.xlsx')
for file in files:
    print(file + 'を読み込みました。')


# In[43]:


wb = load_workbook(file)
ws = wb.active
row_list = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                       min_col=1, max_col=ws.max_column):
    cell_list = [cell.value for cell in row]
    if cell_list[5]=='超保険':
        row_list.append(cell_list) #エクセルのデータをプログラムで読み込める用に処理する
today = datetime.date.today() #今日の日付を取得
date = today.strftime("%Y年%m月吉日") #日付の書式を変更


# In[44]:


#row_list


# In[45]:


p_temp = 'C:/Users/nlpla/Documents/program/3_超保険の更新書状/'
fontsize = 16


# In[46]:


for i in row_list:
    #print(i)
    if i[0]==None:
        break;
    title = str(i[0])
    wb.create_sheet(title=title) #契約者名のシートを作る
    ws_new = wb[title]
    ws_new.column_dimensions['A'].width = '120' #B列の幅を30文字に設定
    
    ws_new.cell(2,1).value = date #日付を記入
    ws_new.cell(2,1).alignment = Alignment(horizontal='right') #右寄せ
    ws_new.cell(2,1).font = Font(size=fontsize)
    
    ws_new.cell(3,1).value = i[1] + '  様' #契約者名を記入
    ws_new.cell(3,1).font = Font(size=fontsize, underline='single')
    
    senders = ['ライフマスター株式会社　松山支店　保険のぽると',
               '〒791-8025　愛媛県松山市衣山1-188　パルティフジ衣山',
              'TEL：089-924-0606　FAX：089-924-0605',
              'e-mail matsuyama-shiten@lifemeister.com'        
                  ]
    
    for num in enumerate(senders):
        ws_new.cell(num[0]+4,1).value = num[1] #送付元を記入
        ws_new.cell(num[0]+4,1).font = Font(size=fontsize)
        ws_new.cell(num[0]+4,1).alignment = Alignment(horizontal='right') #右寄せ
    
    holidays = ['定休日　火曜日']
    
    for holi in enumerate(holidays):
        ws_new.cell(holi[0]+8,1).value = holi[1] #送付元を記入
        ws_new.cell(holi[0]+8,1).font = Font(size=fontsize,color='FF0000')
        ws_new.cell(holi[0]+8,1).alignment = Alignment(horizontal='right') #右寄せ
    
    ws_new.cell(10,1).value = '〜保険契約の更新について(案内)〜' #タイトルを記入
    ws_new.cell(10,1).font = Font(size=fontsize)
    ws_new.cell(10,1).alignment = Alignment(horizontal='center') #中央寄せ
    
    n = str(pd.to_datetime(str(i[9]))).split(' ')[0].split('-')
    #n = str(datetime.datetime(1899, 12, 30) + datetime.timedelta(days=i[9])).split(' ')[0].split('-')
        
    expire = '{0}年{1}月{2}日'.format(n[0], n[1], n[2])    
    sentences = ['いつもお世話になっております。',
                 '{0}様にご契約いただいている東京海上日動の{1}について、満期日が{2}'.format(i[1],i[5],expire),
                  'となっております。',
                 '近日中に保険会社より継続証が送付されますので、ご契約内容をご確認の上、変更事項',
                 '（下記に記載）がございましたら、満期日の1ヶ月前までに「保険のぽると」へご連絡を',
                 'お願い致します。',
                 '更新日（満期日）までに変更事項のご連絡がない場合は、更新前と同様のご契約内容にて',
                 '自動更新されますので、ご了承ください。',
                 '今後ともどうぞよろしくお願い致します。',
                    ]
    
    for ber in enumerate(sentences):
        ws_new.cell(ber[0]+12,1).value = ber[1]
        ws_new.cell(ber[0]+12,1).font = Font(size=fontsize)
    
    image_file3 = p_temp + 'hoken_porto3.png'
    image3 = Image(image_file3)
    image3.width = 900
    image3.height = 540
    ws_new.add_image(image3, 'A21')
        
    ws_new.row_dimensions[48].height = 40
    ws_new.cell(48,1).value = '人生を笑顔で旅するパートナー' #タイトルを記入
    ws_new.cell(48,1).font = Font(color='FF00FF', size=28)
    ws_new.cell(48,1).alignment = Alignment(horizontal='center') #中央寄せ
    
    image_file = p_temp + 'hoken_porto2.png'
    image = Image(image_file)
    image.width = 800
    image.height = 480
    ws_new.add_image(image, 'A49')
    
    messages = ['保険のぽるとは「生きるのに必要なお金を作るための保険ショップ」',
                '「ぽると」は港（porto）を意味しており、',
                ''
               ]
    
    for j in enumerate(messages):
        ws_new.cell(j[0]+71,1).value = j[1]
        ws_new.cell(j[0]+71,1).font = Font(size=fontsize)
        
    ws_new.cell(73,1).value = 'お金の不安を解消して、また新しい人生に送り出したい' #タイトルを記入
    ws_new.cell(73,1).font = Font(size=fontsize, bold=True)
    
    texts = ['',
                'というメッセージを込めております。',
                'これからも{0}様が安心して人生を旅できるように貢献します。'.format(i[1]),
                'お気軽にご連絡・ご来店ください。',               
               ]
    
    for t in enumerate(texts):
        ws_new.cell(t[0]+74,1).value = t[1]
        ws_new.cell(t[0]+74,1).font = Font(size=fontsize)
    
    #ws_new.cell(78,1).value = '保険のぽるとで相談できること！' #タイトルを記入
    #ws_new.cell(78,1).font = Font(size=fontsize)
    #ws_new.cell(78,1).alignment = Alignment(horizontal='center') #中央寄せ
    
    image_file2 = p_temp + 'porto_business.png'
    image2 = Image(image_file2)
    image2.width = 800
    image2.height = 480
    ws_new.add_image(image2, 'A79')
    
    #ws_new.cell(2,2).font = Font(name='書体', color='色', seize='文字サイズ', bold='太字')


# In[47]:


n


# In[48]:


wb.save(str(file).split('.')[0] + '_comp.xlsx')
print('書状を作成しました。')


# In[49]:


str(file)


# 年齢のデータを追加する場合

# In[50]:


'''
更新の際に連絡が必要な事項
住所変更、補償内容の変更、免許証の色の変更など各種変更

住所・電話番号が変わった
（自動車保険がある方）免許証の色が変わった（例：ゴールドからブルーになった等）
（自動車保険がある方）運転者の範囲を変えたい
家族が増えた
保障内容を変えたい
契約者、被保険者を変更したい
他の保険会社と比較したい
口座・クレジットカードを変更したい
'''


# In[ ]:





# In[ ]:




